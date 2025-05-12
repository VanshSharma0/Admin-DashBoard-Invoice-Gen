import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from calendar import monthrange

class JewelrySalesReport:
    def __init__(self):
        self.brand_name = "Murti Jewellers"
        self.product_categories = {
            'Gold': {
                'Rings': {
                    'sub_categories': [
                        'Diamond Rings', 'Gold Rings', 'Antique Rings', 'Bridal Rings',
                        'Engagement Rings', 'Fashion Rings', 'Platinum Rings', 'Designer Rings'
                    ],
                    'unit_cost': {
                        'Diamond Rings': 5000, 'Gold Rings': 3000, 'Antique Rings': 4500,
                        'Bridal Rings': 6000, 'Engagement Rings': 5500, 'Fashion Rings': 2000,
                        'Platinum Rings': 4000, 'Designer Rings': 3500
                    }
                },
                'Necklaces': {
                    'sub_categories': [
                        'Gold Chains', 'Diamond Pendants', 'Antique Necklaces', 'Bridal Sets',
                        'Choker Necklaces', 'Long Chains', 'Pearl Necklaces', 'Designer Necklaces'
                    ],
                    'unit_cost': {
                        'Gold Chains': 4000, 'Diamond Pendants': 6000, 'Antique Necklaces': 5000,
                        'Bridal Sets': 8000, 'Choker Necklaces': 3000, 'Long Chains': 3500,
                        'Pearl Necklaces': 2000, 'Designer Necklaces': 4500
                    }
                },
                'Earrings': {
                    'sub_categories': [
                        'Diamond Studs', 'Gold Hoops', 'Antique Earrings', 'Chandelier Earrings',
                        'Jhumka Earrings', 'Pearl Earrings', 'Designer Earrings', 'Bridal Earrings'
                    ],
                    'unit_cost': {
                        'Diamond Studs': 3000, 'Gold Hoops': 2000, 'Antique Earrings': 2800,
                        'Chandelier Earrings': 3000, 'Jhumka Earrings': 2500, 'Pearl Earrings': 1500,
                        'Designer Earrings': 2200, 'Bridal Earrings': 3500
                    }
                },
                'Bracelets': {
                    'sub_categories': [
                        'Gold Bracelets', 'Diamond Bracelets', 'Antique Bracelets', 'Bridal Bracelets',
                        'Charm Bracelets', 'Pearl Bracelets', 'Designer Bracelets', 'Fashion Bracelets'
                    ],
                    'unit_cost': {
                        'Gold Bracelets': 3500, 'Diamond Bracelets': 4500, 'Antique Bracelets': 4000,
                        'Bridal Bracelets': 5000, 'Charm Bracelets': 2500, 'Pearl Bracelets': 2200,
                        'Designer Bracelets': 3000, 'Fashion Bracelets': 1800
                    }
                },
                'Bangles': {
                    'sub_categories': [
                        'Gold Bangles', 'Diamond Bangles', 'Antique Bangles', 'Bridal Bangles',
                        'Charm Bangles', 'Pearl Bangles', 'Designer Bangles', 'Fashion Bangles'
                    ],
                    'unit_cost': {
                        'Gold Bangles': 3000, 'Diamond Bangles': 5000, 'Antique Bangles': 2500,
                        'Bridal Bangles': 4000, 'Charm Bangles': 2000, 'Pearl Bangles': 1800,
                        'Designer Bangles': 2800, 'Fashion Bangles': 1500
                    }
                }
            },
            'Silver': {
                'Rings': {
                    'sub_categories': [
                        'Silver Rings', 'Antique Silver Rings', 'Fashion Silver Rings',
                        'Designer Silver Rings', 'Bridal Silver Rings', 'Engagement Silver Rings'
                    ],
                    'unit_cost': {
                        'Silver Rings': 1000, 'Antique Silver Rings': 1500, 'Fashion Silver Rings': 1200,
                        'Designer Silver Rings': 1800, 'Bridal Silver Rings': 2000, 'Engagement Silver Rings': 1800
                    }
                },
                'Necklaces': {
                    'sub_categories': [
                        'Silver Chains', 'Antique Silver Necklaces', 'Fashion Silver Necklaces',
                        'Designer Silver Necklaces', 'Pearl Silver Necklaces', 'Bridal Silver Necklaces'
                    ],
                    'unit_cost': {
                        'Silver Chains': 1500, 'Antique Silver Necklaces': 2000, 'Fashion Silver Necklaces': 1800,
                        'Designer Silver Necklaces': 2500, 'Pearl Silver Necklaces': 2200, 'Bridal Silver Necklaces': 3000
                    }
                },
                'Earrings': {
                    'sub_categories': [
                        'Silver Earrings', 'Antique Silver Earrings', 'Fashion Silver Earrings',
                        'Designer Silver Earrings', 'Pearl Silver Earrings', 'Bridal Silver Earrings'
                    ],
                    'unit_cost': {
                        'Silver Earrings': 800, 'Antique Silver Earrings': 1200, 'Fashion Silver Earrings': 1000,
                        'Designer Silver Earrings': 1500, 'Pearl Silver Earrings': 1300, 'Bridal Silver Earrings': 1800
                    }
                },
                'Bracelets': {
                    'sub_categories': [
                        'Silver Bracelets', 'Antique Silver Bracelets', 'Fashion Silver Bracelets',
                        'Designer Silver Bracelets', 'Pearl Silver Bracelets', 'Bridal Silver Bracelets'
                    ],
                    'unit_cost': {
                        'Silver Bracelets': 1200, 'Antique Silver Bracelets': 1800, 'Fashion Silver Bracelets': 1500,
                        'Designer Silver Bracelets': 2000, 'Pearl Silver Bracelets': 1800, 'Bridal Silver Bracelets': 2500
                    }
                },
                'Anklets': {
                    'sub_categories': [
                        'Silver Anklets', 'Antique Silver Anklets', 'Fashion Silver Anklets',
                        'Designer Silver Anklets', 'Pearl Silver Anklets', 'Bridal Silver Anklets'
                    ],
                    'unit_cost': {
                        'Silver Anklets': 800, 'Antique Silver Anklets': 1200, 'Fashion Silver Anklets': 1000,
                        'Designer Silver Anklets': 1500, 'Pearl Silver Anklets': 1300, 'Bridal Silver Anklets': 1800
                    }
                }
            }
        }
        
    def _get_date_list(self, year, month):
        """Generate list of dates for the current month"""
        num_days = monthrange(year, month)[1]
        return [f"{year}-{month:02d}-{day:02d}" for day in range(1, num_days + 1)]

    def create_empty_report(self, month, year):
        """Create an empty DataFrame with the required categories and sub-categories"""
        # Create individual sales entries sheet
        sales_entries = pd.DataFrame({
            'Date': [],
            'Metal Type': [],
            'Category': [],
            'Sub-Category': [],
            'Quantity': [],
            'Unit Price': [],
            'Total Amount': [],
            'Payment Method': [],
            'Customer Name': [],
            'Notes': []
        })
        
        # Create summary sheets for Gold and Silver
        summary_data = {
            'Gold': [],
            'Silver': []
        }
        
        for metal_type, categories in self.product_categories.items():
            for category, details in categories.items():
                for sub_category in details['sub_categories']:
                    summary_data[metal_type].append({
                        'Category': category,
                        'Sub-Category': sub_category,
                        'Total Quantity': 0,
                        'Total Sales': 0.0,
                        'Unit Cost': details['unit_cost'][sub_category],
                        'Total Cost': 0.0,
                        'Profit': 0.0,
                        'Profit Margin': 0.0
                    })
        
        # Create Excel file with multiple sheets
        filename = f'{self.brand_name}_Sales_{month}_{year}.xlsx'
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Sales entries sheet
            sales_entries.to_excel(writer, sheet_name='Sales Entries', index=False)
            
            # Summary sheets
            pd.DataFrame(summary_data['Gold']).to_excel(writer, sheet_name='Gold Summary', index=False)
            pd.DataFrame(summary_data['Silver']).to_excel(writer, sheet_name='Silver Summary', index=False)
            
            # Weekly summary sheet
            weekly_df = pd.DataFrame({
                'Week': [f'Week {i+1}' for i in range(4)],
                'Gold Sales': [0] * 4,
                'Silver Sales': [0] * 4,
                'Total Sales': [0] * 4,
                'Gold Items': [0] * 4,
                'Silver Items': [0] * 4,
                'Total Items': [0] * 4
            })
            weekly_df.to_excel(writer, sheet_name='Weekly Summary', index=False)
            
            # Monthly comparison sheet
            monthly_df = pd.DataFrame({
                'Month': ['January', 'February', 'March', 'April', 'May', 'June',
                         'July', 'August', 'September', 'October', 'November', 'December'],
                'Gold Sales': [0] * 12,
                'Silver Sales': [0] * 12,
                'Total Sales': [0] * 12,
                'Gold Items': [0] * 12,
                'Silver Items': [0] * 12,
                'Total Items': [0] * 12
            })
            monthly_df.to_excel(writer, sheet_name='Monthly Comparison', index=False)
        
        self._format_excel_file(filename, year, month)
        print(f"Empty report created: {filename}")
        return filename

    def _format_excel_file(self, filename, year, month):
        """Format the Excel file with proper styling"""
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        brand_font = Font(name='Arial', size=16, bold=True, color='000000')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add brand name to each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.insert_rows(1, 2)
            ws.merge_cells('A1:J1')
            ws['A1'] = self.brand_name
            ws['A1'].font = brand_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Add date range
            ws['A2'] = f'Report Period: {datetime(year, month, 1).strftime("%B %Y")}'
            ws['A2'].font = Font(bold=True)
            ws.merge_cells('A2:J2')
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Format headers
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Format data cells
            for row in ws.iter_rows(min_row=4):
                for cell in row:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                        if 'Price' in ws[3][cell.column-1].value or 'Amount' in ws[3][cell.column-1].value:
                            cell.number_format = '"₹ "#,##0.00'
                        elif 'Margin' in ws[3][cell.column-1].value:
                            cell.number_format = '0.00"%"'
                    else:
                        cell.alignment = Alignment(horizontal='left')
            
            # Add data validation for Sales Entries sheet
            if sheet_name == 'Sales Entries':
                # Date validation
                dates = self._get_date_list(year, month)
                dv = DataValidation(type="list", formula1=f'"{",".join(dates)}"')
                ws.add_data_validation(dv)
                dv.add(f'A4:A1000')
                
                # Metal Type validation
                metal_types = list(self.product_categories.keys())
                dv = DataValidation(type="list", formula1=f'"{",".join(metal_types)}"')
                ws.add_data_validation(dv)
                dv.add(f'B4:B1000')
                
                # Category validation (dynamic based on metal type)
                for row in range(4, 1000):
                    ws[f'C{row}'] = f'=IF(B{row}="Gold",INDIRECT("Gold_Categories"),INDIRECT("Silver_Categories"))'
                
                # Payment Method validation
                payment_methods = ['Cash', 'Credit Card', 'Debit Card', 'UPI', 'Bank Transfer']
                dv = DataValidation(type="list", formula1=f'"{",".join(payment_methods)}"')
                ws.add_data_validation(dv)
                dv.add(f'H4:H1000')
                
                # Add formulas
                for row in range(4, 1000):
                    # Total Amount = Quantity * Unit Price
                    ws[f'G{row}'] = f'=E{row}*F{row}'
            
            # Add formulas for Summary sheets
            if sheet_name in ['Gold Summary', 'Silver Summary']:
                for row in range(4, 1000):
                    # Total Cost = Total Quantity * Unit Cost
                    ws[f'F{row}'] = f'=C{row}*E{row}'
                    # Profit = Total Sales - Total Cost
                    ws[f'G{row}'] = f'=D{row}-F{row}'
                    # Profit Margin = (Profit / Total Sales) * 100
                    ws[f'H{row}'] = f'=IF(D{row}=0,0,(G{row}/D{row})*100)'
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add named ranges for categories
        ws = wb['Sales Entries']
        gold_categories = list(self.product_categories['Gold'].keys())
        silver_categories = list(self.product_categories['Silver'].keys())
        
        ws['K1'] = ','.join(gold_categories)
        ws['L1'] = ','.join(silver_categories)
        
        wb.define_name('Gold_Categories', f'=Sales_Entries!$K$1')
        wb.define_name('Silver_Categories', f'=Sales_Entries!$L$1')
        
        wb.save(filename)

    def update_sales_data(self, filename):
        """Update the sales data in the Excel file"""
        # Read sales entries
        sales_df = pd.read_excel(filename, sheet_name='Sales Entries')
        
        # Calculate summary data for each metal type
        summary_data = {
            'Gold': [],
            'Silver': []
        }
        
        for metal_type, categories in self.product_categories.items():
            for category, details in categories.items():
                for sub_category in details['sub_categories']:
                    mask = (sales_df['Metal Type'] == metal_type) & (sales_df['Category'] == category) & (sales_df['Sub-Category'] == sub_category)
                    total_quantity = sales_df.loc[mask, 'Quantity'].sum()
                    total_sales = sales_df.loc[mask, 'Total Amount'].sum()
                    unit_cost = details['unit_cost'][sub_category]
                    total_cost = total_quantity * unit_cost
                    profit = total_sales - total_cost
                    profit_margin = (profit / total_sales * 100) if total_sales > 0 else 0
                    
                    summary_data[metal_type].append({
                        'Category': category,
                        'Sub-Category': sub_category,
                        'Total Quantity': total_quantity,
                        'Total Sales': total_sales,
                        'Unit Cost': unit_cost,
                        'Total Cost': total_cost,
                        'Profit': profit,
                        'Profit Margin': profit_margin
                    })
        
        # Calculate weekly and monthly summaries
        sales_df['Date'] = pd.to_datetime(sales_df['Date'])
        sales_df['Week'] = sales_df['Date'].dt.isocalendar().week
        sales_df['Month'] = sales_df['Date'].dt.month
        
        weekly_summary = sales_df.groupby(['Week', 'Metal Type']).agg({
            'Total Amount': 'sum',
            'Quantity': 'sum'
        }).reset_index()
        
        monthly_summary = sales_df.groupby(['Month', 'Metal Type']).agg({
            'Total Amount': 'sum',
            'Quantity': 'sum'
        }).reset_index()
        
        # Create charts
        self._create_sales_charts(summary_data, filename)
        
        # Save updated data
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            sales_df.to_excel(writer, sheet_name='Sales Entries', index=False)
            pd.DataFrame(summary_data['Gold']).to_excel(writer, sheet_name='Gold Summary', index=False)
            pd.DataFrame(summary_data['Silver']).to_excel(writer, sheet_name='Silver Summary', index=False)
            
            # Update weekly summary
            weekly_df = pd.read_excel(filename, sheet_name='Weekly Summary')
            for _, row in weekly_summary.iterrows():
                week_idx = row['Week'] - 1
                if 0 <= week_idx < 4:
                    if row['Metal Type'] == 'Gold':
                        weekly_df.loc[week_idx, 'Gold Sales'] = row['Total Amount']
                        weekly_df.loc[week_idx, 'Gold Items'] = row['Quantity']
                    else:
                        weekly_df.loc[week_idx, 'Silver Sales'] = row['Total Amount']
                        weekly_df.loc[week_idx, 'Silver Items'] = row['Quantity']
                    weekly_df.loc[week_idx, 'Total Sales'] = weekly_df.loc[week_idx, 'Gold Sales'] + weekly_df.loc[week_idx, 'Silver Sales']
                    weekly_df.loc[week_idx, 'Total Items'] = weekly_df.loc[week_idx, 'Gold Items'] + weekly_df.loc[week_idx, 'Silver Items']
            weekly_df.to_excel(writer, sheet_name='Weekly Summary', index=False)
            
            # Update monthly comparison
            monthly_df = pd.read_excel(filename, sheet_name='Monthly Comparison')
            for _, row in monthly_summary.iterrows():
                month_idx = row['Month'] - 1
                if row['Metal Type'] == 'Gold':
                    monthly_df.loc[month_idx, 'Gold Sales'] = row['Total Amount']
                    monthly_df.loc[month_idx, 'Gold Items'] = row['Quantity']
                else:
                    monthly_df.loc[month_idx, 'Silver Sales'] = row['Total Amount']
                    monthly_df.loc[month_idx, 'Silver Items'] = row['Quantity']
                monthly_df.loc[month_idx, 'Total Sales'] = monthly_df.loc[month_idx, 'Gold Sales'] + monthly_df.loc[month_idx, 'Silver Sales']
                monthly_df.loc[month_idx, 'Total Items'] = monthly_df.loc[month_idx, 'Gold Items'] + monthly_df.loc[month_idx, 'Silver Items']
            monthly_df.to_excel(writer, sheet_name='Monthly Comparison', index=False)
        
        self._format_excel_file(filename, sales_df['Date'].dt.year.iloc[0], sales_df['Date'].dt.month.iloc[0])
        print(f"Report updated successfully: {filename}")

    def _create_sales_charts(self, summary_data, filename):
        """Create and save sales charts"""
        # Create directory for charts if it doesn't exist
        charts_dir = 'sales_charts'
        os.makedirs(charts_dir, exist_ok=True)
        
        # Set style
        plt.style.use('seaborn')
        
        # Create charts for each metal type
        for metal_type in ['Gold', 'Silver']:
            df = pd.DataFrame(summary_data[metal_type])
            
            # 1. Sales by Category (Bar Chart)
            plt.figure(figsize=(12, 6))
            category_sales = df.groupby('Category')['Total Sales'].sum()
            sns.barplot(x=category_sales.index, y=category_sales.values)
            plt.title(f'{self.brand_name} - {metal_type} Sales by Category')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(f'{charts_dir}/{metal_type.lower()}_sales_by_category.png')
            plt.close()
            
            # 2. Profit Margin by Sub-Category (Horizontal Bar Chart)
            plt.figure(figsize=(12, 8))
            subcategory_profit = df.groupby('Sub-Category')['Profit Margin'].mean().sort_values()
            sns.barplot(y=subcategory_profit.index, x=subcategory_profit.values)
            plt.title(f'{self.brand_name} - {metal_type} Average Profit Margin by Sub-Category')
            plt.xlabel('Profit Margin (%)')
            plt.tight_layout()
            plt.savefig(f'{charts_dir}/{metal_type.lower()}_profit_margin_by_subcategory.png')
            plt.close()
            
            # 3. Sales Distribution (Pie Chart)
            plt.figure(figsize=(10, 10))
            plt.pie(category_sales, labels=category_sales.index, autopct='%1.1f%%')
            plt.title(f'{self.brand_name} - {metal_type} Sales Distribution by Category')
            plt.tight_layout()
            plt.savefig(f'{charts_dir}/{metal_type.lower()}_sales_distribution.png')
            plt.close()
            
            # 4. Top 10 Sub-Categories by Sales
            plt.figure(figsize=(12, 6))
            top_subcategories = df.groupby('Sub-Category')['Total Sales'].sum().nlargest(10)
            sns.barplot(x=top_subcategories.index, y=top_subcategories.values)
            plt.title(f'{self.brand_name} - Top 10 {metal_type} Sub-Categories by Sales')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(f'{charts_dir}/{metal_type.lower()}_top_subcategories.png')
            plt.close()

def main():
    report = JewelrySalesReport()
    
    # Get current month and year
    current_date = datetime.now()
    month = current_date.month
    year = current_date.year
    
    # Create empty report
    filename = report.create_empty_report(month, year)
    
    print("\nInstructions:")
    print("1. Open the generated Excel file and enter your sales data in the 'Sales Entries' sheet")
    print("2. For each sale, enter:")
    print("   - Select Date from dropdown")
    print("   - Select Metal Type (Gold/Silver) from dropdown")
    print("   - Select Category from dropdown (changes based on Metal Type)")
    print("   - Enter Sub-Category")
    print("   - Enter Quantity and Unit Price")
    print("   - Select Payment Method from dropdown")
    print("   - Add Customer Name and Notes if needed")
    print("3. The Total Amount will be calculated automatically")
    print("4. Save the file")
    print("5. Run this script again to update the report with your data")
    print(f"\nThe report has been created: {filename}")
    print("\nThe report includes:")
    print("- Separate summaries for Gold and Silver items")
    print("- Detailed sales entries with automatic calculations")
    print("- Weekly and monthly comparisons")
    print("- Multiple visual charts in the 'sales_charts' directory")

if __name__ == "__main__":
    main() 